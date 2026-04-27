"""
IFRS S2 Climate Disclosure Gap Analyser
========================================
Upload a company's annual report (PDF) and receive a comprehensive gap analysis
benchmarked against all IFRS S2 disclosure requirements.
"""

import streamlit as st
import pdfplumber
import anthropic
import pandas as pd
import json
import io
import time
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# IFRS S2 REQUIREMENTS LIBRARY
# ─────────────────────────────────────────────────────────────────────────────

IFRS_S2_REQUIREMENTS = [
    # ── GOVERNANCE ──────────────────────────────────────────────────────────
    {
        "category": "Governance",
        "name": "Governance Body(s)/Individual(s) Identification",
        "description": "Identify the specific governance body(s) (e.g., board, committee) or individual(s) responsible for oversight of climate-related risks and opportunities.",
        "applicable_scope": "Entity-specific governance structure (IFRS S2, para 6(a))",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Responsibilities in Mandates/Policies",
        "description": "Disclose how responsibilities for climate-related risks and opportunities are formally reflected in the terms of reference, mandates, role descriptions, and other related policies.",
        "applicable_scope": "Oversight body/individual mandates & policies (IFRS S2, para 6(a)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Oversight Body Skills & Competencies Determination",
        "description": "Disclose how the governance body(s) or individual(s) determines whether appropriate skills and competencies are available or will be developed to oversee strategies designed to respond to climate-related risks and opportunities.",
        "applicable_scope": "Oversight body/individual skills assessment (IFRS S2, para 6(a)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Frequency & Nature of Climate Information to Oversight Body",
        "description": "Disclose how and how often the governance body(s) or individual(s) is informed about climate-related risks and opportunities.",
        "applicable_scope": "Oversight body/individual information flow (IFRS S2, para 6(a)(iii))",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Consideration of Climate in Oversight (Strategy, Transactions, Risk)",
        "description": "Disclose how the governance body(s) or individual(s) considers climate-related risks and opportunities when overseeing the entity's strategy, decisions on major transactions, risk management processes and related policies, including consideration of associated trade-offs.",
        "applicable_scope": "Oversight body/individual strategic & risk oversight (IFRS S2, para 6(a)(iv))",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Oversight of Climate Targets & Monitoring",
        "description": "Disclose how the governance body(s) or individual(s) oversees the setting of climate-related targets and monitors progress towards achieving them.",
        "applicable_scope": "Oversight body/individual; Climate-related targets (IFRS S2, para 6(a)(v), 33-36)",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Linkage of Climate Performance Metrics to Remuneration",
        "description": "Disclose whether and how related performance metrics for climate-related targets are included in remuneration policies.",
        "applicable_scope": "Remuneration policies; Executive management (IFRS S2, para 6(a)(v), 29(g))",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Governance",
        "name": "Management's Role Delegation",
        "description": "Disclose whether management's role in climate oversight is delegated to a specific management-level position or committee, and how oversight is exercised over that delegated role or committee.",
        "applicable_scope": "Management structure & delegation (IFRS S2, para 6(b)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Governance",
        "name": "Management's Controls & Procedures",
        "description": "Disclose whether management uses specific controls and procedures to support climate oversight and, if so, how these are integrated with other internal functions (e.g., risk, finance).",
        "applicable_scope": "Management processes; Internal functions integration (IFRS S2, para 6(b)(ii))",
        "type": "Qualitative",
    },
    # ── STRATEGY ────────────────────────────────────────────────────────────
    {
        "category": "Strategy",
        "name": "Description of Climate-Related Risks & Opportunities",
        "description": "Describe the specific climate-related risks and opportunities that could reasonably be expected to affect the entity's prospects over short, medium, and long term.",
        "applicable_scope": "Entity prospects (IFRS S2, para 10(a))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Risk Classification (Physical vs. Transition)",
        "description": "Explain whether each identified climate-related risk is classified as a physical risk (acute or chronic) or a transition risk (policy, legal, tech, market, reputation).",
        "applicable_scope": "Identified climate risks (IFRS S2, para 10(b))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Time Horizon of Risk/Opportunity Effects",
        "description": "Specify over which time horizons—short, medium, or long term—the effects of each identified climate-related risk and opportunity are expected to occur.",
        "applicable_scope": "Identified climate risks & opportunities (IFRS S2, para 10(c))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Definition of Time Horizons",
        "description": "Explain how the entity defines 'short term', 'medium term', and 'long term' for its disclosures, and how these link to strategic planning and capital allocation.",
        "applicable_scope": "Entity planning horizons; Strategic decision-making (IFRS S2, para 10(d))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Consideration of Industry-Based Disclosure Topics",
        "description": "In identifying risks/opportunities, refer to and consider the applicability of industry-based disclosure topics in the Industry-based Guidance on Implementing IFRS S2.",
        "applicable_scope": "Entity prospects; Industry identification; Industry-based Guidance (IFRS S2, para 12)",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Effects on Business Model & Value Chain (Current & Anticipated)",
        "description": "Describe the current and anticipated future effects of identified climate-related risks and opportunities on the entity's business model and its value chain.",
        "applicable_scope": "Business model; Value chain (IFRS S2, para 13(a))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Concentration of Risks/Opportunities in Business Model/Value Chain",
        "description": "Describe where within the business model and value chain the identified climate-related risks and opportunities are concentrated (e.g., specific geographies, facilities, asset types).",
        "applicable_scope": "Business model; Value chain; Geographical areas; Facilities; Assets (IFRS S2, para 13(b))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Response Strategy & Decision-Making",
        "description": "Disclose how the entity has responded, and plans to respond, to identified climate-related risks and opportunities within its strategy and decision-making processes.",
        "applicable_scope": "Entity strategy & decision-making (IFRS S2, para 14(a))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Plans to Achieve Climate Targets",
        "description": "Disclose specifically how the entity plans to achieve any climate-related targets it has set and any targets it is required to meet by law or regulation.",
        "applicable_scope": "Climate-related targets (entity-set or regulatory) (IFRS S2, para 14(a), 14(a)(v))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Changes to Business Model & Resource Allocation",
        "description": "Disclose current and anticipated changes to the business model, including resource allocation decisions, to address climate issues (e.g., plans for carbon-intensive operations, R&D spend, capex, acquisitions/divestments).",
        "applicable_scope": "Business model; Resource allocation; Capital expenditure; R&D; M&A (IFRS S2, para 14(a)(i))",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Strategy",
        "name": "Direct Mitigation & Adaptation Efforts",
        "description": "Disclose current and anticipated direct efforts to mitigate climate risks or adapt to climate change (e.g., changes in production processes, facility relocation, workforce adjustments).",
        "applicable_scope": "Operations; Facilities; Workforce; Products (IFRS S2, para 14(a)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Indirect Mitigation & Adaptation Efforts",
        "description": "Disclose current and anticipated indirect efforts to mitigate or adapt, involving external parties (e.g., working with customers and supply chain partners).",
        "applicable_scope": "Customers; Supply chain (IFRS S2, para 14(a)(iii))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Climate-Related Transition Plan Details",
        "description": "If the entity has a climate-related transition plan, disclose details including key assumptions and critical dependencies on which the plan relies.",
        "applicable_scope": "Transition plan (if exists); Strategy (IFRS S2, para 14(a)(iv))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Resourcing of Strategic Plans",
        "description": "Provide information about how the entity is currently resourcing, and plans to resource, the strategic activities disclosed in response to climate risks and opportunities.",
        "applicable_scope": "Strategic response activities; Financial & other resources (IFRS S2, para 14(b))",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Strategy",
        "name": "Progress on Previously Disclosed Plans",
        "description": "Provide quantitative and qualitative information regarding the progress made on implementing plans previously disclosed.",
        "applicable_scope": "Previously disclosed strategic plans (IFRS S2, para 14(c))",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Current Financial Effects",
        "description": "Disclose how identified climate-related risks and opportunities have affected the entity's financial position, financial performance, and cash flows for the reporting period.",
        "applicable_scope": "Financial statements (current period) (IFRS S2, para 15(a), 16(a))",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Significant Adjustment Risks in Next Period",
        "description": "Identify those climate risks/opportunities for which there is a significant risk of a material adjustment within the next annual reporting period to the carrying amounts of assets and liabilities.",
        "applicable_scope": "Financial statements (carrying amounts); Next reporting period (IFRS S2, para 16(b))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Anticipated Effects on Financial Position",
        "description": "Disclose how the entity expects its financial position to change over the short, medium, and long term, given its climate strategy.",
        "applicable_scope": "Financial position (future); Investment/disposal plans; Funding sources (IFRS S2, para 16(c))",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Anticipated Effects on Financial Performance & Cash Flows",
        "description": "Disclose how the entity expects its financial performance and cash flows to change over short, medium, and long term given its climate strategy.",
        "applicable_scope": "Financial performance (future); Cash flows (future) (IFRS S2, para 16(d))",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Qualitative Information (if Quantitative Not Provided)",
        "description": "If quantitative information on financial effects is omitted, explain why and provide qualitative information, identifying affected financial statement lines.",
        "applicable_scope": "Financial effects where quantitative data is omitted (IFRS S2, para 21)",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Climate Resilience Assessment",
        "description": "Disclose the entity's assessment of its climate resilience using climate-related scenario analysis to understand resilience to climate changes, developments, and uncertainties.",
        "applicable_scope": "Entity strategy & business model; Scenario analysis (IFRS S2, para 22, 22(a))",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Strategy",
        "name": "Implications of Resilience Assessment",
        "description": "Disclose the implications of the resilience assessment for the entity's strategy and business model, including how the entity would need to respond to effects identified.",
        "applicable_scope": "Entity strategy & business model (IFRS S2, para 22(a)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Significant Uncertainties in Resilience Assessment",
        "description": "Disclose the significant areas of uncertainty considered in the entity's assessment of its climate resilience.",
        "applicable_scope": "Resilience assessment; Uncertainty analysis (IFRS S2, para 22(a)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Capacity to Adjust/Adapt Strategy & Business Model",
        "description": "Disclose the entity's capacity to adjust or adapt its strategy and business model to climate change, including availability of financial resources, ability to redeploy assets, and effect of climate investments.",
        "applicable_scope": "Strategy & business model adaptability; Financial resources; Assets (IFRS S2, para 22(a)(iii))",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Strategy",
        "name": "Climate-Related Scenario Analysis Methodology",
        "description": "Disclose how and when the climate-related scenario analysis used for resilience assessment was carried out.",
        "applicable_scope": "Climate resilience assessment; Scenario analysis (IFRS S2, para 22(b))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Scenario Analysis Inputs (Scenarios, Range, Type, Rationale, Horizon, Scope)",
        "description": "Provide info on scenario analysis inputs: specific scenarios used and their sources; range diversity; associated risk type; alignment with international agreements; relevance rationale; time horizons; operational scope.",
        "applicable_scope": "Scenario analysis methodology & inputs (IFRS S2, para 22(b)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Scenario Analysis Assumptions (Policy, Macro, Regional, Energy, Tech)",
        "description": "Disclose key assumptions made in the scenario analysis regarding climate policies, macroeconomic trends, national/regional variables, energy usage/mix, and technology developments.",
        "applicable_scope": "Scenario analysis methodology & assumptions (IFRS S2, para 22(b)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Strategy",
        "name": "Scenario Analysis Timing",
        "description": "Disclose the reporting period in which the climate-related scenario analysis was carried out.",
        "applicable_scope": "Scenario analysis methodology; Reporting cycle (IFRS S2, para 22(b)(iii), B18)",
        "type": "Qualitative",
    },
    # ── RISK MANAGEMENT ──────────────────────────────────────────────────────
    {
        "category": "Risk Management",
        "name": "Risk Identification, Assessment, Prioritisation & Monitoring Process",
        "description": "Disclose the processes and related policies the entity uses to identify, assess, prioritise, and monitor climate-related risks.",
        "applicable_scope": "Risk management process for climate risks (IFRS S2, para 25(a))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Inputs & Parameters for Risk Process",
        "description": "Disclose the inputs and parameters used within the climate risk processes (e.g., data sources, scope of operations covered).",
        "applicable_scope": "Risk management process inputs (IFRS S2, para 25(a)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Use of Scenario Analysis in Risk Identification",
        "description": "Disclose whether and how the entity uses climate-related scenario analysis to inform its identification of climate-related risks.",
        "applicable_scope": "Risk identification process; Scenario analysis (IFRS S2, para 25(a)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Risk Assessment Approach (Nature, Likelihood, Magnitude)",
        "description": "Disclose how the entity assesses the nature, likelihood, and magnitude of the effects of identified climate risks.",
        "applicable_scope": "Risk assessment process criteria (IFRS S2, para 25(a)(iii))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Prioritisation of Climate Risks",
        "description": "Disclose whether and how the entity prioritises climate-related risks relative to other types of risk within its overall risk framework.",
        "applicable_scope": "Risk prioritisation process (IFRS S2, para 25(a)(iv))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Monitoring of Climate Risks",
        "description": "Disclose how the entity monitors climate-related risks over time.",
        "applicable_scope": "Risk monitoring process (IFRS S2, para 25(a)(v))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Changes in Risk Processes",
        "description": "Disclose whether and how the entity has changed its climate risk management processes compared with the previous reporting period.",
        "applicable_scope": "Risk management process evolution (IFRS S2, para 25(a)(vi))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Opportunity Identification, Assessment, Prioritisation & Monitoring Process",
        "description": "Disclose the processes the entity uses to identify, assess, prioritise, and monitor climate-related opportunities, including whether scenario analysis is used.",
        "applicable_scope": "Opportunity management process; Scenario analysis (IFRS S2, para 25(b))",
        "type": "Qualitative",
    },
    {
        "category": "Risk Management",
        "name": "Integration with Overall Risk Management",
        "description": "Disclose the extent to which, and how, the processes for managing climate risks and opportunities are integrated into the entity's overall risk management process.",
        "applicable_scope": "Overall risk management process integration (IFRS S2, para 25(c))",
        "type": "Qualitative",
    },
    # ── METRICS & TARGETS ────────────────────────────────────────────────────
    {
        "category": "Metrics & Targets",
        "name": "Cross-Industry Metric Categories",
        "description": "Disclose information relevant to the seven specified cross-industry metric categories: GHG emissions, transition risk, physical risk, opportunities, capital deployment, carbon price, and remuneration.",
        "applicable_scope": "Cross-industry applicability (IFRS S2, para 28(a), 29)",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Greenhouse Gas (GHG) Emissions (Scope 1)",
        "description": "Disclose absolute gross Scope 1 GHG emissions generated during the reporting period, expressed as metric tonnes of CO2 equivalent.",
        "applicable_scope": "Scope 1 sources (owned/controlled) (IFRS S2, para 29(a)(i)(1))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Greenhouse Gas (GHG) Emissions (Scope 2)",
        "description": "Disclose absolute gross Scope 2 GHG emissions generated during the reporting period in metric tonnes of CO2 equivalent, using a location-based method.",
        "applicable_scope": "Scope 2 sources (purchased energy) (IFRS S2, para 29(a)(i)(2), 29(a)(v))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Greenhouse Gas (GHG) Emissions (Scope 3)",
        "description": "Disclose absolute gross Scope 3 GHG emissions generated during the reporting period in metric tonnes of CO2 equivalent.",
        "applicable_scope": "Scope 3 sources (value chain) (IFRS S2, para 29(a)(i)(3))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Measurement Standard Used",
        "description": "Measure GHG emissions using the GHG Protocol Corporate Accounting and Reporting Standard (2004) and disclose the method used.",
        "applicable_scope": "GHG emissions measurement methodology (IFRS S2, para 29(a)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Measurement Approach, Inputs, Assumptions",
        "description": "Disclose the specific approach used to measure GHG emissions, the inputs and assumptions used, the reason for choosing them, and any changes made during the period.",
        "applicable_scope": "GHG emissions measurement details (IFRS S2, para 29(a)(iii))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Disaggregation of Scope 1 & 2 Emissions",
        "description": "Disaggregate Scope 1 and Scope 2 emissions between the consolidated accounting group and other investees.",
        "applicable_scope": "Scope 1 & 2 emissions allocation (IFRS S2, para 29(a)(iv))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Scope 2 Contractual Instruments Information",
        "description": "Alongside location-based Scope 2 emissions, provide information about any contractual instruments (e.g., PPAs, RECs) to understand the entity's Scope 2 emissions profile.",
        "applicable_scope": "Scope 2 emissions; Contractual instruments (IFRS S2, para 29(a)(v), B30-B31)",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Scope 3 Categories Included",
        "description": "Disclose which categories are included within the entity's measure of Scope 3 emissions, per the 15 categories in the GHG Protocol Corporate Value Chain Standard.",
        "applicable_scope": "Scope 3 emissions measurement scope (IFRS S2, para 29(a)(vi)(1))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Financed Emissions (Specific Sectors)",
        "description": "If activities include asset management, commercial banking, or insurance, disclose additional information about Category 15 / financed emissions.",
        "applicable_scope": "Scope 3 Category 15; Specific financial sectors (IFRS S2, para 29(a)(vi)(2), B58-B63)",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Climate-Related Transition Risks Exposure (Assets/Activities)",
        "description": "Disclose the amount and percentage of assets or business activities considered vulnerable to climate-related transition risks.",
        "applicable_scope": "Assets; Business activities; Transition risks (IFRS S2, para 29(b))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Climate-Related Physical Risks Exposure (Assets/Activities)",
        "description": "Disclose the amount and percentage of assets or business activities considered vulnerable to climate-related physical risks.",
        "applicable_scope": "Assets; Business activities; Physical risks (IFRS S2, para 29(c))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Climate-Related Opportunities Alignment (Assets/Activities)",
        "description": "Disclose the amount and percentage of assets or business activities considered aligned with climate-related opportunities.",
        "applicable_scope": "Assets; Business activities; Climate opportunities (IFRS S2, para 29(d))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Capital Deployment Towards Climate",
        "description": "Disclose the amount of capital expenditure, financing, or investment deployed towards climate-related risks and opportunities during the period.",
        "applicable_scope": "Capital expenditure; Financing; Investment (IFRS S2, para 29(e))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Internal Carbon Pricing Application",
        "description": "Explain whether and how the entity applies an internal carbon price in its decision-making processes.",
        "applicable_scope": "Internal decision-making; Internal carbon price use (IFRS S2, para 29(f)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Internal Carbon Price Value",
        "description": "Disclose the price(s) used for each metric tonne of GHG emissions when assessing the costs of emissions internally.",
        "applicable_scope": "Internal carbon price value(s) (IFRS S2, para 29(f)(ii))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Climate Considerations in Executive Remuneration (Description)",
        "description": "Provide a description of whether and how climate-related considerations are factored into executive remuneration policies and decisions.",
        "applicable_scope": "Executive remuneration policies (IFRS S2, para 29(g)(i))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Climate-Linked Executive Remuneration (Percentage)",
        "description": "Disclose the percentage of executive management remuneration recognised in the current period that is linked to climate-related considerations.",
        "applicable_scope": "Executive remuneration (current period) (IFRS S2, para 29(g)(ii))",
        "type": "Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Industry-Based Metrics",
        "description": "Disclose industry-based metrics associated with the entity's business models, activities, or other industry features, referring to the Industry-based Guidance on Implementing IFRS S2.",
        "applicable_scope": "Industry-specific activities/business models; Industry-based Guidance (IFRS S2, para 28(b), 32)",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Climate-Related Targets (Set or Required)",
        "description": "Disclose the quantitative and qualitative climate-related targets the entity has set, and any targets required by law or regulation, including GHG targets.",
        "applicable_scope": "Entity-set targets; Regulatory targets (IFRS S2, para 28(c), 33)",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Metric Used",
        "description": "Disclose the specific metric used to set the target and track progress.",
        "applicable_scope": "Specific target metric (IFRS S2, para 33(a))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Objective",
        "description": "Disclose the objective of the target (e.g., climate change mitigation, adaptation, conformance with science-based initiatives).",
        "applicable_scope": "Specific target objective (IFRS S2, para 33(b))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Scope of Application",
        "description": "Disclose the part of the entity to which the target applies (e.g., the entire entity, a specific business unit, a specific geographical region).",
        "applicable_scope": "Specific target scope (IFRS S2, para 33(c))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Time Period",
        "description": "Disclose the period over which the target applies.",
        "applicable_scope": "Specific target timeframe (IFRS S2, para 33(d))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Base Period",
        "description": "Disclose the base period from which progress towards the target is measured.",
        "applicable_scope": "Specific target baseline (IFRS S2, para 33(e))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Milestones & Interim Targets",
        "description": "Disclose any milestones and interim targets set on the path to achieving the main target.",
        "applicable_scope": "Specific target pathway (IFRS S2, para 33(f))",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Type (Absolute vs. Intensity)",
        "description": "If the target is quantitative, disclose whether it is an absolute target (total amount) or an intensity target (ratio relative to a business metric).",
        "applicable_scope": "Specific quantitative target type (IFRS S2, para 33(g))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Details: Link to International Agreements",
        "description": "Disclose how the latest international agreement on climate change (e.g., Paris Agreement) has informed the target.",
        "applicable_scope": "Specific target context; International climate agreements (IFRS S2, para 33(h))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Approach to Setting & Reviewing Targets",
        "description": "Disclose information about the entity's approach to setting and reviewing each target, and how it monitors progress against each target.",
        "applicable_scope": "Target setting & review process (IFRS S2, para 34)",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Validation",
        "description": "Disclose whether the target and the methodology for setting it have been validated by a third party.",
        "applicable_scope": "Specific target validation status (IFRS S2, para 34(a))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Review Process",
        "description": "Disclose the entity's internal processes for reviewing the target periodically.",
        "applicable_scope": "Specific target review mechanisms (IFRS S2, para 34(b))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Metrics for Monitoring Target Progress",
        "description": "Disclose the specific metrics used to monitor progress towards reaching the target.",
        "applicable_scope": "Specific target monitoring metrics (IFRS S2, para 34(c))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Target Revisions",
        "description": "Disclose any revisions made to the target and provide an explanation for those revisions.",
        "applicable_scope": "Specific target changes (IFRS S2, para 34(d))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Performance Against Targets",
        "description": "Disclose information about the entity's performance against each climate-related target, including an analysis of trends or changes over time.",
        "applicable_scope": "Specific target performance (IFRS S2, para 35)",
        "type": "Quantitative / Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Target Details: Gases Covered",
        "description": "For each GHG emissions target, specify which greenhouse gases (e.g., CO2 only, all Kyoto Protocol gases) are covered.",
        "applicable_scope": "Specific GHG target gas coverage (IFRS S2, para 36(a))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Target Details: Scope Coverage (1, 2, 3)",
        "description": "For each GHG emissions target, specify whether Scope 1, Scope 2, or Scope 3 emissions are covered.",
        "applicable_scope": "Specific GHG target scope coverage (IFRS S2, para 36(b))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Target Details: Gross vs. Net",
        "description": "For each GHG emissions target, specify whether it is a gross emissions target or a net emissions target.",
        "applicable_scope": "Specific GHG target type (gross/net) (IFRS S2, para 36(c), B68-B69)",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Target Details: Sectoral Decarbonisation Approach",
        "description": "For each GHG emissions target, specify whether the target was derived using a sectoral decarbonisation approach.",
        "applicable_scope": "Specific GHG target derivation method (IFRS S2, para 36(d))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "GHG Target Details: Planned Use of Carbon Credits",
        "description": "For any net GHG emissions target, disclose the entity's planned use of carbon credits to offset emissions to achieve the target.",
        "applicable_scope": "Specific net GHG target; Carbon credits strategy (IFRS S2, para 36(e), B70)",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Carbon Credit Details: Reliance",
        "description": "Disclose the extent to which, and how, achieving any net GHG emissions target relies on the planned use of carbon credits.",
        "applicable_scope": "Specific net GHG target reliance on credits (IFRS S2, para 36(e)(i))",
        "type": "Qualitative / Quantitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Carbon Credit Details: Verification Scheme",
        "description": "Disclose which third-party scheme(s) will verify or certify the carbon credits the entity plans to use.",
        "applicable_scope": "Planned carbon credit verification (IFRS S2, para 36(e)(ii))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Carbon Credit Details: Type (Nature-based/Tech, Reduction/Removal)",
        "description": "Disclose the type of carbon credit planned, including whether nature-based or technological, and whether achieved through carbon reduction or removal.",
        "applicable_scope": "Planned carbon credit type (IFRS S2, para 36(e)(iii))",
        "type": "Qualitative",
    },
    {
        "category": "Metrics & Targets",
        "name": "Carbon Credit Details: Credibility & Integrity Factors",
        "description": "Disclose any other factors necessary for users to understand the credibility and integrity of the carbon credits the entity plans to use.",
        "applicable_scope": "Planned carbon credit quality factors (IFRS S2, para 36(e)(iv))",
        "type": "Qualitative",
    },
]

CATEGORIES = ["Governance", "Strategy", "Risk Management", "Metrics & Targets"]

MATERIALITY_SCALE = ["Very High", "High", "Moderate", "Low", "Very Low"]

# ─────────────────────────────────────────────────────────────────────────────
# PDF TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf_text(uploaded_file) -> str:
    """Extract all text from uploaded PDF, return as a single string."""
    text_parts = []
    with pdfplumber.open(uploaded_file) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text()
            if page_text:
                text_parts.append(f"[PAGE {page_num}]\n{page_text}")
    return "\n\n".join(text_parts)


# ─────────────────────────────────────────────────────────────────────────────
# CLAUDE API ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an expert in sustainability reporting, specialising in IFRS S1 and IFRS S2 standards issued by the International Sustainability Standards Board (ISSB). Your objective is to perform a rigorous gap analysis of a company's climate-related disclosures against the IFRS S2 standard.

Work in analytical and meticulous mode. Provide concise but evidence-based rationales. When page numbers are referenced in the text (e.g., [PAGE 42]), use them for citations."""

def build_analysis_prompt(company_name: str, report_text: str, requirements: list, category: str) -> str:
    req_list = "\n".join([
        f'{i+1}. **{r["name"]}**\n   Description: {r["description"]}\n   Applicable Scope: {r["applicable_scope"]}'
        for i, r in enumerate(requirements)
    ])

    # Truncate report text if very long (keep ~120k chars ≈ ~30k tokens)
    MAX_CHARS = 120_000
    truncated = report_text[:MAX_CHARS]
    if len(report_text) > MAX_CHARS:
        truncated += "\n\n[NOTE: Report text truncated for processing. Analysis based on available content.]"

    return f"""Analyse the following annual report for **{company_name}** against the IFRS S2 **{category}** disclosure requirements listed below.

For EACH requirement, return a JSON object with these exact fields:
- "requirement_name": string (exact name from the list)
- "fulfillment_status": "Yes" | "Partial" | "No"
- "disclosure_summary": string (2-4 sentences summarising what the company disclosed or why it is missing. Be specific and quote or paraphrase actual content from the report.)
- "page_numbers": string (comma-separated page numbers where evidence was found, e.g. "22, 23, 58". Leave blank if No.)
- "materiality_level": one of: "Very High" | "High" | "Moderate" | "Low" | "Very Low"
- "materiality_justification": string (1-2 sentences explaining why this requirement is material given the company's industry and sector)
- "recommended_enhancements": string (2-3 specific, actionable recommendations aligned with IFRS S2 for improving this disclosure)

Return a JSON array containing one object per requirement. Output ONLY the JSON array — no markdown fences, no preamble, no commentary.

---
IFRS S2 {category.upper()} REQUIREMENTS TO ANALYSE:
{req_list}

---
ANNUAL REPORT TEXT FOR {company_name.upper()}:
{truncated}
"""


def analyse_category(
    client: anthropic.Anthropic,
    company_name: str,
    report_text: str,
    category: str,
    requirements: list,
) -> list[dict]:
    """Call Claude to analyse one category of IFRS S2 requirements."""
    prompt = build_analysis_prompt(company_name, report_text, requirements, category)

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=8000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = response.content[0].text.strip()

    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    try:
        results = json.loads(raw)
    except json.JSONDecodeError:
        # Attempt partial recovery
        results = []
        for req in requirements:
            results.append({
                "requirement_name": req["name"],
                "fulfillment_status": "No",
                "disclosure_summary": "Analysis failed for this requirement. Please re-run.",
                "page_numbers": "",
                "materiality_level": "Moderate",
                "materiality_justification": "Unable to assess.",
                "recommended_enhancements": "Re-run analysis.",
            })

    # Enrich results with static metadata from requirements
    req_lookup = {r["name"]: r for r in requirements}
    for item in results:
        req = req_lookup.get(item.get("requirement_name"), {})
        item["category"] = category
        item["applicable_scope"] = req.get("applicable_scope", "")
        item["type"] = req.get("type", "")

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

# Colour palette
COLOURS = {
    "header_bg":       "1F3864",   # dark navy
    "header_font":     "FFFFFF",
    "cat_governance":  "D6E4F0",   # pale blue
    "cat_strategy":    "D5F5E3",   # pale green
    "cat_risk":        "FEF9E7",   # pale yellow
    "cat_metrics":     "FADBD8",   # pale pink
    "yes_fill":        "C8E6C9",
    "partial_fill":    "FFF9C4",
    "no_fill":         "FFCDD2",
    "mat_very_high":   "B71C1C",
    "mat_high":        "E53935",
    "mat_moderate":    "FB8C00",
    "mat_low":         "43A047",
    "mat_very_low":    "1E88E5",
}

CAT_COLOURS = {
    "Governance":       COLOURS["cat_governance"],
    "Strategy":         COLOURS["cat_strategy"],
    "Risk Management":  COLOURS["cat_risk"],
    "Metrics & Targets": COLOURS["cat_metrics"],
}

STATUS_COLOURS = {
    "Yes":     COLOURS["yes_fill"],
    "Partial": COLOURS["partial_fill"],
    "No":      COLOURS["no_fill"],
}

MATERIALITY_COLOURS = {
    "Very High": COLOURS["mat_very_high"],
    "High":      COLOURS["mat_high"],
    "Moderate":  COLOURS["mat_moderate"],
    "Low":       COLOURS["mat_low"],
    "Very Low":  COLOURS["mat_very_low"],
}

COLUMNS = [
    "IFRS S2 Disclosure Requirement",
    "Fulfillment Status",
    "Disclosure Summary",
    "Page Number(s)",
    "Materiality (Sector-Relevant)",
    "Recommended Enhancements",
    "Applicable Scope",
    "Type (Qualitative/Quantitative)",
    "Category",
]

COL_WIDTHS = [45, 14, 70, 14, 30, 70, 55, 25, 18]


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(company_name: str, results: list[dict]) -> bytes:
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    total = len(results)
    yes_count = sum(1 for r in results if r.get("fulfillment_status") == "Yes")
    partial_count = sum(1 for r in results if r.get("fulfillment_status") == "Partial")
    no_count = sum(1 for r in results if r.get("fulfillment_status") == "No")

    title_font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=COLOURS["header_bg"])

    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = f"IFRS S2 Gap Analysis — {company_name}"
    ws_sum["A1"].font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
    ws_sum["A1"].fill = header_fill
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    ws_sum["A3"] = "Metric"
    ws_sum["B3"] = "Count"
    ws_sum["C3"] = "% of Total"
    for col in ["A3", "B3", "C3"]:
        ws_sum[col].font = Font(name="Calibri", bold=True, color="FFFFFF")
        ws_sum[col].fill = header_fill
        ws_sum[col].alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Total Requirements", total, "100%"),
        ("✅  Fulfilled (Yes)", yes_count, f"{yes_count/total*100:.1f}%" if total else "0%"),
        ("⚠️  Partial", partial_count, f"{partial_count/total*100:.1f}%" if total else "0%"),
        ("❌  Not Fulfilled (No)", no_count, f"{no_count/total*100:.1f}%" if total else "0%"),
    ]
    for i, (label, count, pct) in enumerate(summary_rows, start=4):
        ws_sum[f"A{i}"] = label
        ws_sum[f"B{i}"] = count
        ws_sum[f"C{i}"] = pct
        for col in [f"A{i}", f"B{i}", f"C{i}"]:
            ws_sum[col].alignment = Alignment(horizontal="center")
            ws_sum[col].border = thin_border()

    # Per-category breakdown
    ws_sum["A9"] = "Category Breakdown"
    ws_sum["A9"].font = Font(name="Calibri", bold=True, size=13)

    ws_sum["A10"] = "Category"
    ws_sum["B10"] = "Yes"
    ws_sum["C10"] = "Partial"
    ws_sum["D10"] = "No"
    for col in ["A10", "B10", "C10", "D10"]:
        ws_sum[col].font = Font(name="Calibri", bold=True, color="FFFFFF")
        ws_sum[col].fill = header_fill
        ws_sum[col].alignment = Alignment(horizontal="center")

    for i, cat in enumerate(CATEGORIES, start=11):
        cat_results = [r for r in results if r.get("category") == cat]
        y = sum(1 for r in cat_results if r.get("fulfillment_status") == "Yes")
        p = sum(1 for r in cat_results if r.get("fulfillment_status") == "Partial")
        n = sum(1 for r in cat_results if r.get("fulfillment_status") == "No")
        ws_sum[f"A{i}"] = cat
        ws_sum[f"B{i}"] = y
        ws_sum[f"C{i}"] = p
        ws_sum[f"D{i}"] = n
        fill = PatternFill("solid", fgColor=CAT_COLOURS.get(cat, "FFFFFF"))
        for col in [f"A{i}", f"B{i}", f"C{i}", f"D{i}"]:
            ws_sum[col].fill = fill
            ws_sum[col].alignment = Alignment(horizontal="center")
            ws_sum[col].border = thin_border()

    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 12

    # ── Detail sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Benchmark Analysis")

    # Column headers
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", bold=True, size=11, color=COLOURS["header_font"])
        cell.fill = PatternFill("solid", fgColor=COLOURS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32

    row_idx = 2
    for result in results:
        cat = result.get("category", "")
        status = result.get("fulfillment_status", "No")
        materiality = result.get("materiality_level", "Moderate")
        materiality_text = f"{materiality} — {result.get('materiality_justification', '')}"

        row_data = [
            result.get("requirement_name", ""),
            status,
            result.get("disclosure_summary", ""),
            result.get("page_numbers", ""),
            materiality_text,
            result.get("recommended_enhancements", ""),
            result.get("applicable_scope", ""),
            result.get("type", ""),
            cat,
        ]

        base_fill = PatternFill("solid", fgColor=CAT_COLOURS.get(cat, "FFFFFF"))
        status_fill = PatternFill("solid", fgColor=STATUS_COLOURS.get(status, "FFFFFF"))
        mat_font_color = MATERIALITY_COLOURS.get(materiality, "000000")

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()

            if col_idx == 2:  # Fulfillment Status
                cell.fill = status_fill
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 5:  # Materiality
                cell.fill = base_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color=mat_font_color)
            else:
                cell.fill = base_fill

        ws.row_dimensions[row_idx].height = 80
        row_idx += 1

    # Freeze panes
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="IFRS S2 Gap Analyser",
    page_icon="🌍",
    layout="wide",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1F3864;
        margin-bottom: 0.2rem;
    }
    .sub-title {
        font-size: 1rem;
        color: #555;
        margin-bottom: 1.5rem;
    }
    .stat-box {
        background: #f0f4fa;
        border-radius: 10px;
        padding: 16px 20px;
        text-align: center;
        border: 1px solid #dce4f0;
    }
    .stat-label { font-size: 0.85rem; color: #666; }
    .stat-value { font-size: 2rem; font-weight: 700; color: #1F3864; }
    .status-yes     { color: #2e7d32; font-weight: 700; }
    .status-partial { color: #e65100; font-weight: 700; }
    .status-no      { color: #c62828; font-weight: 700; }
    .section-header {
        font-size: 1.2rem;
        font-weight: 600;
        color: #1F3864;
        border-left: 4px solid #1F3864;
        padding-left: 8px;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">🌍 IFRS S2 Climate Disclosure Gap Analyser</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Upload a company\'s annual report and receive an AI-powered gap analysis benchmarked against all IFRS S2 disclosure requirements.</div>', unsafe_allow_html=True)

st.divider()

# ── Sidebar — Configuration ──────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Configuration")

    api_key = st.text_input(
        "Anthropic API Key",
        type="password",
        help="Your Anthropic API key. Get one at console.anthropic.com",
    )

    company_name = st.text_input(
        "Company Name",
        placeholder="e.g. SLP Resources Berhad",
        help="The name of the company whose annual report you are analysing.",
    )

    st.divider()
    st.markdown("**About IFRS S2**")
    st.markdown(
        "IFRS S2 (Climate-related Disclosures) is an ISSB standard requiring entities "
        "to disclose climate-related risks and opportunities across four pillars: "
        "Governance, Strategy, Risk Management, and Metrics & Targets."
    )
    st.markdown(f"**{len(IFRS_S2_REQUIREMENTS)} requirements** across **4 categories** will be assessed.")

# ── Main — Upload ─────────────────────────────────────────────────────────────
col1, col2 = st.columns([2, 1])

with col1:
    st.markdown('<div class="section-header">📄 Upload Annual Report (PDF)</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Drag and drop or click to upload",
        type=["pdf"],
        help="Upload the company's Integrated Annual Report or Sustainability Report in PDF format.",
    )

with col2:
    st.markdown('<div class="section-header">📋 Requirements Coverage</div>', unsafe_allow_html=True)
    for cat in CATEGORIES:
        count = sum(1 for r in IFRS_S2_REQUIREMENTS if r["category"] == cat)
        st.metric(cat, f"{count} requirements")

# ── Analysis trigger ──────────────────────────────────────────────────────────
st.divider()

run_analysis = st.button(
    "🔍 Run IFRS S2 Analysis",
    type="primary",
    disabled=not (api_key and company_name and uploaded_file),
    use_container_width=True,
)

if not api_key:
    st.info("👈 Enter your Anthropic API key in the sidebar to get started.")
elif not company_name:
    st.info("👈 Enter the company name in the sidebar.")
elif not uploaded_file:
    st.info("⬆️ Upload the company's annual report PDF above.")

# ── Run analysis ──────────────────────────────────────────────────────────────
if run_analysis:
    try:
        client = anthropic.Anthropic(api_key=api_key)

        # Step 1: Extract PDF
        with st.status("Extracting text from PDF...", expanded=True) as status:
            st.write(f"📄 Processing: {uploaded_file.name}")
            report_text = extract_pdf_text(uploaded_file)
            word_count = len(report_text.split())
            st.write(f"✅ Extracted ~{word_count:,} words from the report")
            status.update(label="PDF extraction complete", state="complete")

        # Step 2: Analyse each category
        all_results = []
        progress_bar = st.progress(0, text="Starting analysis…")
        status_text = st.empty()

        for i, category in enumerate(CATEGORIES):
            cat_reqs = [r for r in IFRS_S2_REQUIREMENTS if r["category"] == category]
            progress = i / len(CATEGORIES)
            progress_bar.progress(progress, text=f"Analysing {category} ({len(cat_reqs)} requirements)…")
            status_text.markdown(f"**🔄 Analysing: {category}** — {len(cat_reqs)} requirements")

            with st.spinner(f"Running AI analysis for {category}…"):
                results = analyse_category(client, company_name, report_text, category, cat_reqs)
                all_results.extend(results)

            progress_bar.progress((i + 1) / len(CATEGORIES), text=f"✅ {category} complete")
            time.sleep(0.3)

        progress_bar.progress(1.0, text="Analysis complete!")
        status_text.empty()

        # ── Results ───────────────────────────────────────────────────────────
        st.success(f"✅ Analysis complete! {len(all_results)} requirements assessed for {company_name}.")

        # Summary stats
        yes_count = sum(1 for r in all_results if r.get("fulfillment_status") == "Yes")
        partial_count = sum(1 for r in all_results if r.get("fulfillment_status") == "Partial")
        no_count = sum(1 for r in all_results if r.get("fulfillment_status") == "No")

        st.markdown("### 📊 Summary")
        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f'<div class="stat-box"><div class="stat-label">Total Requirements</div><div class="stat-value">{len(all_results)}</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="stat-box"><div class="stat-label">✅ Fulfilled</div><div class="stat-value status-yes">{yes_count}</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="stat-box"><div class="stat-label">⚠️ Partial</div><div class="stat-value status-partial">{partial_count}</div></div>', unsafe_allow_html=True)
        c4.markdown(f'<div class="stat-box"><div class="stat-label">❌ Not Fulfilled</div><div class="stat-value status-no">{no_count}</div></div>', unsafe_allow_html=True)

        st.divider()

        # Detailed results by category
        st.markdown("### 📋 Detailed Results")

        for category in CATEGORIES:
            cat_results = [r for r in all_results if r.get("category") == category]
            y = sum(1 for r in cat_results if r.get("fulfillment_status") == "Yes")
            p = sum(1 for r in cat_results if r.get("fulfillment_status") == "Partial")
            n = sum(1 for r in cat_results if r.get("fulfillment_status") == "No")

            with st.expander(f"**{category}** — {y} Yes / {p} Partial / {n} No", expanded=(category == "Governance")):
                for result in cat_results:
                    status = result.get("fulfillment_status", "No")
                    status_icon = {"Yes": "✅", "Partial": "⚠️", "No": "❌"}.get(status, "❓")
                    mat = result.get("materiality_level", "Moderate")
                    mat_icon = {"Very High": "🔴", "High": "🟠", "Moderate": "🟡", "Low": "🟢", "Very Low": "🔵"}.get(mat, "⚪")

                    st.markdown(f"**{status_icon} {result.get('requirement_name', '')}** &nbsp; {mat_icon} *{mat} Materiality*")
                    st.markdown(f"> {result.get('disclosure_summary', '')}")
                    if result.get("page_numbers"):
                        st.markdown(f"📄 Pages: `{result.get('page_numbers')}`")
                    if result.get("recommended_enhancements"):
                        st.markdown(f"💡 **Recommended:** {result.get('recommended_enhancements')}")
                    st.markdown("---")

        # ── Download ──────────────────────────────────────────────────────────
        st.markdown("### 💾 Download Results")
        excel_bytes = build_excel(company_name, all_results)

        st.download_button(
            label="⬇️ Download Excel Report (.xlsx)",
            data=excel_bytes,
            file_name=f"IFRS_S2_Gap_Analysis_{company_name.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

        st.caption("The Excel file contains a Summary sheet with statistics and a detailed Benchmark Analysis sheet with all findings, colour-coded by fulfillment status and materiality.")

    except anthropic.AuthenticationError:
        st.error("❌ Invalid API key. Please check your Anthropic API key in the sidebar.")
    except anthropic.RateLimitError:
        st.error("❌ Rate limit reached. Please wait a moment and try again.")
    except Exception as e:
        st.error(f"❌ An error occurred: {str(e)}")
        st.exception(e)
